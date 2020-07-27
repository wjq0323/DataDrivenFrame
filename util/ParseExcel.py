#读取Excel文件

import openpyxl
from openpyxl.styles import  Border,Side,Font
import time

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None)
        #颜色对应的RGB值
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        #将Excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        #根据sheet名获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return  sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,index):
        #根据index获取sheet
        try:
            sheetname = self.workbook.get_sheet_names()[index]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        #获取最大行号
        return sheet.max_row

    def getColsNumber(self,sheet):
        # 获取最大列号
        return sheet.max_column

    def getStartRowsNumber(self,sheet):
        #开始的行号
        return sheet.min_row

    def getStartColsNumber(self,sheet):
        #开始的列号
        return sheet.min_col

    def getRow(self,sheet,rowNo):
        #获取某一行 返回的是这一行所有的数据内容组成的tuple 下标从1开始
        try:
            return sheet.rows[rowNo-1] #去掉表头
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        try:
            return sheet.columns[colNo - 1] #去掉第一列序号
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        #根据单元格所在的位置获取单元格的值 比如sheet.cell(row =1,col =1).value 表示第一行第一列的值
        #coordinate表示坐标 比如A1
        if coordinate is not None:
            try:
                return sheet.cell(coordinate = coordinate).vaule
            except Exception as e:
                return e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("insufficient Coordinates of cell!")

    def getCellOfObject(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        #根据单元格所在的位置获取单元格对象
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                return e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo , column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("insufficient Coordinates of cell!")

    def writeCell(self, sheet,content, coordinate = None, rowNo = None, colsNo = None,style = None):
        #向单元格写入数据 style是字体的颜色
        if coordinate is not None:
            try:
                 sheet.cell(coordinate = coordinate).value = content
                 if style is not None:
                     sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                 self.workbook.save(self.excelFile)
            except Exception as e:
                return e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("insufficient Coordinates of cell!")

    def writeCellCurrentTime(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        #写入当前的时间 下标从1开始
        now = int(time.time()) #时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                 sheet.cell(coordinate = coordinate).value = currentTime
                 self.workbook.save(self.excelFile)
            except Exception as e:
                return e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("insufficient Coordinates of cell!")

if __name__ == '__main__':
    pe = ParseExcel()
    pe.loadWorkBook("../testData/126邮箱联系人.xlsx")
    print("通过名称获取sheet对象的名字",pe.getSheetByName(u"联系人").title)
    print("通过序号获取sheet对象的名字", pe.getSheetByIndex(0).title)
    sheet = pe.getSheetByIndex(0)
    print(type(sheet))
    print(pe.getRowsNumber(sheet))
    print(pe.getColsNumber(sheet))
    rows = pe.getRow(sheet,1) #获取第一行
    for i in rows:
        print(i.value)
    print(pe.getCellOfValue(sheet,rowNo=1,colsNo=1))
    pe.writeCell(sheet,'wjq',rowNo=10,colsNo=10)
    pe.writeCellCurrentTime(sheet,rowNo=10,colsNo=11)

